"""
Created on Mon Apr  7 08:57:43 2025

@author: BaeHun
"""

# openxl 라이브러리 가져오기
import openpyxl

# 엑셀 파일 불러오기
file_path = 'character.xlsx' # 파일이 있는 경로 또는 같은 폴더에 있을 경우
wb = openpyxl.load_workbook(file_path)
print(wb.sheetnames)
sheet1 = wb['Sheet1']
print(sheet1['A1'].value)
wb.create_sheet('Sheet2')
print(wb.sheetnames)

# 기존 시트 이름으로 접근
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

# 시트 이름 변경
sheet1.title = '캐릭터 명단'
sheet2.title = '인기도 조사'

# 변경된 시트 이름 목록 출력
print(wb.sheetnames)

# 변경 사항 저장
wb.save('characters_renamed.xlsx')

sheet2['B1'] = '인기도 조사결과'
print(sheet2['B1'].value)

copysheet = wb.copy_worksheet(sheet2)
print("복제 전 시트 목록:", wb.sheetnames)

# 복제 시트 이름 변경
copysheet.title = 'copy'

# 시트 목록 확인
print("복제 후 시트 목록:", wb.sheetnames)

# 복제된 시트의 B1 셀 값 출력
print("copy 시트의 B1 셀 값:", copysheet['B1'].value)

# 변경 내용 저장 (선택사항)
wb.save('characters_modified.xlsx')

del wb['인기도 조사']
print(wb.sheetnames)

wb.save('complete.xlsx')

